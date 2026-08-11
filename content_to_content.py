import streamlit as st
import pandas as pd
import openpyxl
import re
import io
import zipfile
import warnings

warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

# 1. PAGE CONFIG & TRUCK ART STYLING
st.set_page_config(page_title="C 2 C | Goods Carrier", layout="wide", page_icon="🚚")

st.markdown("""<style>
@import url('https://fonts.googleapis.com/css2?family=Yatra+One&family=Poppins:wght@400;600;700;800&display=swap');

html, body, [class*="css"] {
    font-family: 'Poppins', sans-serif;
    background-color: #FAFAFA;
}

/* Indian Truck Art Header Container */
.truck-header-card {
    background: #FFFDF5;
    border: 3px solid #1E293B;
    border-radius: 14px;
    padding: 24px;
    margin-bottom: 24px;
    box-shadow: 6px 6px 0px #1E293B;
    position: relative;
    overflow: hidden;
}

/* Top Truck Art Chevron Bar */
.truck-art-border {
    height: 10px;
    background: repeating-linear-gradient(
        45deg,
        #FF9933,
        #FF9933 15px,
        #E0115F 15px,
        #E0115F 30px,
        #FFD700 30px,
        #FFD700 45px,
        #0F766E 45px,
        #0F766E 60px
    );
    margin: -24px -24px 20px -24px;
    border-bottom: 2px solid #1E293B;
}

.brand-wrapper {
    display: flex;
    align-items: center;
    gap: 28px;
}

.brand-title-text {
    font-family: 'Yatra One', cursive;
    font-size: 3.8rem;
    color: #1E293B;
    line-height: 1;
    margin: 0;
    text-shadow: 2px 2px 0px #FF9933;
    letter-spacing: 2px;
}

/* Horn OK Please Style Badge */
.horn-ok-badge {
    display: inline-flex;
    align-items: center;
    gap: 8px;
    background-color: #FFD700;
    color: #1E293B;
    font-family: 'Poppins', sans-serif;
    font-weight: 800;
    font-size: 0.8rem;
    text-transform: uppercase;
    letter-spacing: 2px;
    padding: 6px 16px;
    border-radius: 6px;
    border: 2px solid #1E293B;
    box-shadow: 3px 3px 0px #1E293B;
    margin-top: 8px;
}

/* Truck Bumper Button */
div.stButton > button:first-child {
    background: #FF9933;
    color: #1E293B;
    font-family: 'Poppins', sans-serif;
    font-weight: 800;
    font-size: 1.05rem;
    text-transform: uppercase;
    letter-spacing: 2px;
    border: 3px solid #1E293B;
    padding: 0.85rem 2.5rem;
    border-radius: 10px;
    box-shadow: 5px 5px 0px #1E293B;
    transition: all 0.15s ease-in-out;
}

div.stButton > button:first-child:hover {
    background: #FFD700;
    color: #1E293B;
    box-shadow: 2px 2px 0px #1E293B;
    transform: translate(3px, 3px);
}
</style>""", unsafe_allow_html=True)

# 2. HELPER FUNCTIONS
def map_attribute_header(raw_attr):
    raw = str(raw_attr).lower().strip()
    if 'display' in raw or 'title' in raw or raw == 'style name': return 'productdisplayname'
    if 'list view' in raw: return 'listviewname'
    if 'product detail' in raw or 'description' in raw: return 'productdetails'
    if 'size' in raw and 'fit' in raw: return 'sizeandfitdescription'
    if 'material' in raw and 'care' in raw: return 'materialcaredescription'
    if 'colour' in raw or 'color' in raw: return 'colour'
    if raw == 'patterns': return 'pattern'
    if raw == 'fashion trends': return 'maintrend'
    return re.sub(r'[^a-z0-9]', '', raw)

def clean_id(val):
    try: return str(int(float(str(val).strip())))
    except: return str(val).strip()

def normalize_col(name):
    name = str(name).strip().lower()
    return re.sub(r'[^a-z0-9]', '', name)

# 3. TRUCK ART HEADER HTML
HEADER_HTML = """<div class="truck-header-card">
<div class="truck-art-border"></div>
<div class="brand-wrapper">
<svg width="130" height="85" viewBox="0 0 170 100" fill="none" xmlns="http://www.w3.org/2000/svg">
<path d="M15 70 H55 V25 H35 L20 42 L15 70 Z" fill="#FF9933" stroke="#1E293B" stroke-width="2.5"/>
<path d="M22 45 H35 V32 L26 42 Z" fill="#FFFDF5" stroke="#1E293B" stroke-width="2"/>
<rect x="55" y="18" width="100" height="52" rx="4" fill="#0F766E" stroke="#1E293B" stroke-width="2.5"/>
<path d="M60 24 L70 32 L80 24 L90 32 L100 24 L110 32 L120 24 L130 32 L140 24 L150 32" stroke="#E0115F" stroke-width="2.5"/>
<path d="M60 64 L70 56 L80 64 L90 56 L100 64 L110 56 L120 64 L130 56 L140 64 L150 56" stroke="#FFD700" stroke-width="2.5"/>
<rect x="68" y="34" width="22" height="20" rx="3" fill="#FF9933" stroke="#FFFDF5" stroke-width="2"/>
<rect x="94" y="34" width="22" height="20" rx="3" fill="#FFD700" stroke="#FFFDF5" stroke-width="2"/>
<rect x="120" y="34" width="22" height="20" rx="3" fill="#E0115F" stroke="#FFFDF5" stroke-width="2"/>
<path d="M105 44 C102 39 98 41 105 35 C112 41 108 39 105 44 Z" fill="#E0115F"/>
<circle cx="35" cy="74" r="10" fill="#1E293B" stroke="#FFD700" stroke-width="2.5"/>
<circle cx="35" cy="74" r="4" fill="#FFFDF5"/>
<circle cx="85" cy="74" r="10" fill="#1E293B" stroke="#FFD700" stroke-width="2.5"/>
<circle cx="85" cy="74" r="4" fill="#FFFDF5"/>
<circle cx="135" cy="74" r="10" fill="#1E293B" stroke="#FFD700" stroke-width="2.5"/>
<circle cx="135" cy="74" r="4" fill="#FFFDF5"/>
<path d="M5 88 H165" stroke="#FF9933" stroke-width="3" stroke-dasharray="8 5" stroke-linecap="round"/>
</svg>
<div>
<div class="brand-title-text">C 2 C</div>
<div class="horn-ok-badge">🪷 MĀL GĀDI • CATALOG CARRIER</div>
</div>
</div>
</div>"""

st.markdown(HEADER_HTML, unsafe_allow_html=True)

# 4. ANIMATED DRIVING TRUCK HTML
ANIMATED_TRUCK_HTML = """<div style="background: #FFFDF5; border: 3px solid #1E293B; border-radius: 12px; padding: 20px; text-align: center; margin: 20px 0; box-shadow: 5px 5px 0px #1E293B;">
<style>
@keyframes driveBounce { 0% { transform: translateY(0px); } 50% { transform: translateY(-3px); } 100% { transform: translateY(0px); } }
@keyframes rotateSpokes { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
@keyframes fillCargoDrop { 0% { opacity: 0; transform: translateY(-18px) scale(0.6); } 40% { opacity: 1; transform: translateY(0px) scale(1); } 100% { opacity: 1; transform: translateY(0px) scale(1); } }
@keyframes roadDash { 0% { stroke-dashoffset: 0; } 100% { stroke-dashoffset: -24; } }
.truck-chassis { animation: driveBounce 0.4s infinite ease-in-out; }
.wheel-spoke { transform-box: fill-box; transform-origin: center; animation: rotateSpokes 0.5s infinite linear; }
.road-line-anim { animation: roadDash 0.25s infinite linear; }
.cargo-box-1 { animation: fillCargoDrop 1.2s infinite ease-out; }
.cargo-box-2 { animation: fillCargoDrop 1.2s infinite ease-out 0.4s; }
.cargo-box-3 { animation: fillCargoDrop 1.2s infinite ease-out 0.8s; }
</style>
<svg width="240" height="110" viewBox="0 0 200 100" fill="none" xmlns="http://www.w3.org/2000/svg">
<path d="M10 82 H190" stroke="#FF9933" stroke-width="3.5" stroke-dasharray="8 6" class="road-line-anim" stroke-linecap="round"/>
<g class="truck-chassis">
<path d="M18 68 H58 V32 H38 L24 45 L18 68 Z" fill="#FF9933" stroke="#1E293B" stroke-width="2"/>
<path d="M26 47 H38 V36 L30 45 Z" fill="#FFFDF5" stroke="#1E293B" stroke-width="1.5"/>
<rect x="58" y="26" width="115" height="42" rx="4" fill="#0F766E" stroke="#1E293B" stroke-width="2"/>
<path d="M62 31 L72 37 L82 31 L92 37 L102 31 L112 37 L122 31 L132 37 L142 31 L152 37 L162 31" stroke="#E0115F" stroke-width="2"/>
<path d="M62 63 L72 57 L82 63 L92 57 L102 63 L112 57 L122 63 L132 57 L142 63 L152 57 L162 63" stroke="#FFD700" stroke-width="2"/>
<g class="cargo-box-1">
<rect x="68" y="38" width="24" height="18" rx="3" fill="#FF9933" stroke="#FFFDF5" stroke-width="1.5"/>
<path d="M68 47 H92" stroke="#FFFDF5" stroke-width="1"/>
</g>
<g class="cargo-box-2">
<rect x="100" y="38" width="24" height="18" rx="3" fill="#FFD700" stroke="#FFFDF5" stroke-width="1.5"/>
<path d="M100 47 H124" stroke="#FFFDF5" stroke-width="1"/>
</g>
<g class="cargo-box-3">
<rect x="132" y="38" width="24" height="18" rx="3" fill="#E0115F" stroke="#FFFDF5" stroke-width="1.5"/>
<path d="M132 47 H156" stroke="#FFFDF5" stroke-width="1"/>
</g>
<g transform="translate(36, 70)">
<circle cx="0" cy="0" r="9" fill="#1E293B" stroke="#FFD700" stroke-width="2.5"/>
<circle cx="0" cy="0" r="3.5" fill="#FFFDF5"/>
<line x1="-7" y1="0" x2="7" y2="0" stroke="#FFFDF5" stroke-width="1.5" class="wheel-spoke"/>
<line x1="0" y1="-7" x2="0" y2="7" stroke="#FFFDF5" stroke-width="1.5" class="wheel-spoke"/>
</g>
<g transform="translate(86, 70)">
<circle cx="0" cy="0" r="9" fill="#1E293B" stroke="#FFD700" stroke-width="2.5"/>
<circle cx="0" cy="0" r="3.5" fill="#FFFDF5"/>
<line x1="-7" y1="0" x2="7" y2="0" stroke="#FFFDF5" stroke-width="1.5" class="wheel-spoke"/>
<line x1="0" y1="-7" x2="0" y2="7" stroke="#FFFDF5" stroke-width="1.5" class="wheel-spoke"/>
</g>
<g transform="translate(142, 70)">
<circle cx="0" cy="0" r="9" fill="#1E293B" stroke="#FFD700" stroke-width="2.5"/>
<circle cx="0" cy="0" r="3.5" fill="#FFFDF5"/>
<line x1="-7" y1="0" x2="7" y2="0" stroke="#FFFDF5" stroke-width="1.5" class="wheel-spoke"/>
<line x1="0" y1="-7" x2="0" y2="7" stroke="#FFFDF5" stroke-width="1.5" class="wheel-spoke"/>
</g>
</g>
</svg>
<div style="font-family: 'Poppins', sans-serif; font-weight: 800; color: #1E293B; margin-top: 8px; text-transform: uppercase; letter-spacing: 2px; font-size: 0.9rem;">
🚚 CHALO! LOADING CATALOG PAYLOAD & FILLING TRUCK...
</div>
</div>"""

# --- FILE UPLOADERS ---
col1, col2 = st.columns(2)

with col1:
    template_files = st.file_uploader(
        "1. Target Myntra Templates (.xlsx)", 
        type=["xlsx"], 
        accept_multiple_files=True,
        key="tmpl_up"
    )

with col2:
    seller_files = st.file_uploader(
        "2. Source Seller Files (.xlsx, .csv)", 
        type=["xlsx", "csv"], 
        accept_multiple_files=True,
        key="sell_up"
    )

st.markdown("---")

# --- EXECUTION ENGINE ---
if st.button("🚚💨 CHALO! RUN C 2 C MAPPING") and template_files and seller_files:
    anim_placeholder = st.empty()
    anim_placeholder.markdown(ANIMATED_TRUCK_HTML, unsafe_allow_html=True)
    
    master_style_dict = {}
    
    # 5. PARSE SELLER FILES
    for uploaded_seller in seller_files:
        try:
            if uploaded_seller.name.endswith('.csv'):
                df = pd.read_csv(uploaded_seller)
                style_col = next((c for c in df.columns if normalize_col(c) in ['styleid', 'styleids', 'style', 'id']), df.columns[0])
                for _, row in df.iterrows():
                    sid = clean_id(row[style_col])
                    if sid and sid != 'nan':
                        if sid not in master_style_dict: master_style_dict[sid] = {}
                        for col in df.columns:
                            if col != style_col and pd.notna(row[col]):
                                master_style_dict[sid][map_attribute_header(col)] = str(row[col]).strip()
                continue

            xls = pd.ExcelFile(uploaded_seller)
            sheet_name = 'Style Sheet' if 'Style Sheet' in xls.sheet_names else ('Styles' if 'Styles' in xls.sheet_names else xls.sheet_names[0])
            df_raw = pd.read_excel(uploaded_seller, sheet_name=sheet_name, header=None)
            
            is_complex_header = False
            for r in range(min(4, len(df_raw))):
                row_str = " ".join([str(x).lower() for x in df_raw.iloc[r].dropna().values])
                if 'field names' in row_str or 'current inputs' in row_str or 'correct inputs' in row_str or 'new inputs' in row_str:
                    is_complex_header = True
                    break

            if is_complex_header:
                start_row = 4 if len(df_raw) > 4 else 2
                for r in range(start_row, len(df_raw)):
                    row = df_raw.iloc[r]
                    raw_sid = row[0]
                    if pd.isna(raw_sid): continue
                    sid = clean_id(raw_sid)
                    if not sid or sid == 'nan': continue
                    
                    if sid not in master_style_dict: master_style_dict[sid] = {}
                    
                    if 10 < len(row) and pd.notna(row[10]):
                        attr_name = str(row[10]).strip()
                        val_curr_spec = row[11] if 11 < len(row) else None
                        val_new_spec = row[12] if 12 < len(row) else None
                        
                        target_val = None
                        if pd.notna(val_new_spec) and str(val_new_spec).strip().lower() not in ['nan', '']:
                            target_val = str(val_new_spec).strip()
                        elif pd.notna(val_curr_spec) and str(val_curr_spec).strip().lower() not in ['nan', '']:
                            target_val = str(val_curr_spec).strip()
                            
                        if attr_name and target_val:
                            master_style_dict[sid][map_attribute_header(attr_name)] = target_val

                    paired_cols = [(4, 5, 4), (6, 7, 6), (8, 9, 8), (13, 14, 13), (15, 16, 15)]
                    for pair_curr, pair_new, header_idx in paired_cols:
                        if header_idx < df_raw.shape[1]:
                            h_val = df_raw.iloc[2, header_idx] if pd.notna(df_raw.iloc[2, header_idx]) else df_raw.iloc[1, header_idx]
                            if pd.notna(h_val):
                                attr_header = str(h_val).strip()
                                v_new = row[pair_new] if pair_new < len(row) else None
                                v_curr = row[pair_curr] if pair_curr < len(row) else None
                                
                                final_v = None
                                if pd.notna(v_new) and str(v_new).strip().lower() not in ['nan', '']:
                                    final_v = str(v_new).strip()
                                elif pd.notna(v_curr) and str(v_curr).strip().lower() not in ['nan', '']:
                                    final_v = str(v_curr).strip()
                                    
                                if attr_header and final_v:
                                    master_style_dict[sid][map_attribute_header(attr_header)] = final_v

            else:
                df_simple = pd.read_excel(uploaded_seller, sheet_name=sheet_name)
                style_col = next((c for c in df_simple.columns if normalize_col(c) in ['styleid', 'styleids', 'style', 'id']), df_simple.columns[0])
                
                for _, row in df_simple.iterrows():
                    sid = clean_id(row[style_col])
                    if not sid or sid == 'nan': continue
                    if sid not in master_style_dict: master_style_dict[sid] = {}
                    
                    for col in df_simple.columns:
                        if col != style_col and pd.notna(row[col]):
                            val = str(row[col]).strip()
                            if val and val.lower() != 'nan':
                                master_style_dict[sid][map_attribute_header(col)] = val

        except Exception as e:
            st.error(f"Error parsing {uploaded_seller.name}: {e}")

    # 6. INJECT INTO TEMPLATES & AUDIT LOG
    zip_buffer = io.BytesIO()
    total_mapped = 0
    total_branded = 0
    audit_records = []
    
    with zipfile.ZipFile(zip_buffer, "a", zipfile.ZIP_DEFLATED, False) as zip_file:
        for uploaded_template in template_files:
            wb = openpyxl.load_workbook(uploaded_template)
            mapped_count = 0
            brand_injection_count = 0

            for sheet_name in wb.sheetnames:
                if sheet_name == 'masterdata': continue
                ws = wb[sheet_name]
                
                col_map = {normalize_col(cell.value): idx for idx, cell in enumerate(ws[1], 1) if cell.value}
                target_style_col = col_map.get('styleid')
                brand_col_idx = col_map.get('brand')
                
                if target_style_col:
                    for row in range(2, ws.max_row + 1):
                        raw_sid = ws.cell(row=row, column=target_style_col).value
                        if not raw_sid: continue
                        sid = clean_id(raw_sid)
                        
                        if sid in master_style_dict:
                            source_data = master_style_dict[sid]
                            for mapped_attr, val in source_data.items():
                                if mapped_attr in col_map:
                                    final_val = str(val).strip()
                                    
                                    if mapped_attr == 'productdisplayname' and brand_col_idx:
                                        target_brand = str(ws.cell(row=row, column=brand_col_idx).value).strip()
                                        if target_brand and target_brand.lower() not in ['none', 'nan', '']:
                                            if target_brand.lower() not in final_val.lower():
                                                final_val = f"{target_brand} {final_val}"
                                                brand_injection_count += 1
                                    
                                    ws.cell(row=row, column=col_map[mapped_attr]).value = final_val
                                    mapped_count += 1
                                    
                                    orig_col_header = ws.cell(row=1, column=col_map[mapped_attr]).value
                                    audit_records.append({
                                        "Template File": uploaded_template.name,
                                        "Sheet Name": sheet_name,
                                        "Style ID": sid,
                                        "Attribute Mapped": orig_col_header,
                                        "Updated Value": final_val
                                    })
            
            total_mapped += mapped_count
            total_branded += brand_injection_count
            
            wb_buffer = io.BytesIO()
            wb.save(wb_buffer)
            zip_file.writestr(f"C2C_Mapped_{uploaded_template.name}", wb_buffer.getvalue())

    # Clear animation container
    anim_placeholder.empty()

    # 7. RESULTS & AUDIT TABLE DISPLAY
    st.success(f"✅ Extracted updates for **{len(master_style_dict)}** unique style IDs from seller files.")
    st.info(f"🎯 **C 2 C Mapping Complete!** Successfully filled **{total_mapped}** attribute cells into catalog template(s).")
    if total_branded > 0:
        st.warning(f"🛡️ Auto-injected Brand Name into **{total_branded}** titles.")

    if audit_records:
        df_audit = pd.DataFrame(audit_records)
        with st.expander("📊 Detailed Mapping Audit Log (Line-by-Line)", expanded=True):
            st.dataframe(df_audit, use_container_width=True)

    if len(template_files) == 1:
        wb_buffer.seek(0)
        st.download_button(
            label="💾 Download Updated C 2 C Template",
            data=wb_buffer.getvalue(),
            file_name=f"C2C_Mapped_{template_files[0].name}",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.download_button(
            label="📦 Download All C 2 C Templates (.zip)",
            data=zip_buffer.getvalue(),
            file_name="C2C_Mapped_Templates.zip",
            mime="application/zip"
        )
